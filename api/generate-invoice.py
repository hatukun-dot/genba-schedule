import json
import base64
import os
import copy
import calendar
from io import BytesIO
from http.server import BaseHTTPRequestHandler
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '請求書テンプレ.xlsx')
SANEI = "サンエイ株式会社営統事業部建設課"


def copy_sheet_styles(ws_src, ws_dst):
    # 列幅
    for col, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col].width = dim.width
        ws_dst.column_dimensions[col].hidden = dim.hidden
    # 行高
    for row, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row].height = dim.height
        ws_dst.row_dimensions[row].hidden = dim.hidden
    # セルスタイル
    for row in ws_src.iter_rows():
        for src_cell in row:
            dst_cell = ws_dst.cell(row=src_cell.row, column=src_cell.column)
            if src_cell.has_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format


def fmt_dates(dates):
    sd = sorted(set(dates))
    parts = []
    i = 0
    while i < len(sd):
        j = i
        while j + 1 < len(sd):
            c = datetime.strptime(sd[j], "%Y-%m-%d")
            n = datetime.strptime(sd[j + 1], "%Y-%m-%d")
            if (n - c).days == 1:
                j += 1
            else:
                break
        s = datetime.strptime(sd[i], "%Y-%m-%d")
        e2 = datetime.strptime(sd[j], "%Y-%m-%d")
        parts.append(f"{s.month}/{s.day}" if i == j else f"{s.month}/{s.day}〜{e2.month}/{e2.day}")
        i = j + 1
    return ".".join(parts)


def generate_invoice(body):
    year = body["year"]
    month_index0 = body["monthIndex0"]
    period_month = month_index0 + 1
    targets = body["targets"]
    events = body["events"]
    projects_map = {p["id"]: p for p in body.get("projects", [])}
    tasks_map = {t["id"]: t for t in body.get("tasks", [])}
    managers_map = {m["id"]: m for m in body.get("managers", [])}
    all_bt = body.get("allBillingTargets", [])

    with open(TEMPLATE_PATH, "rb") as f:
        template_bytes = f.read()

    wb_main = load_workbook(BytesIO(template_bytes))
    wb_main.create_sheet("INDEX", 0)
    ws_idx = wb_main["INDEX"]
    ws_idx["A1"] = "請求書一覧"
    ws_idx["A1"].font = Font(bold=True, size=14)
    ws_idx.cell(2, 1, "請求先").font = Font(bold=True)
    ws_idx.cell(2, 2, "出力タイプ").font = Font(bold=True)
    ws_idx.cell(2, 3, "締め日").font = Font(bold=True)
    ws_idx.column_dimensions["A"].width = 35
    ws_idx.column_dimensions["B"].width = 20
    ws_idx.column_dimensions["C"].width = 12

    index_row = 3
    added_sheets = []

    for target in targets:
        if target.get("outputType") == "出力しない":
            continue

        if target.get("closingType") == "20日締め":
            pm = period_month - 1 if period_month > 1 else 12
            py = year if period_month > 1 else year - 1
            ps = f"{py}-{pm:02d}-21"
            pe = f"{year}-{period_month:02d}-20"
            cd = 20
        else:
            ld = calendar.monthrange(year, period_month)[1]
            ps = f"{year}-{period_month:02d}-01"
            pe = f"{year}-{period_month:02d}-{ld:02d}"
            cd = ld

        rtids = {target["id"]}
        for bt in all_bt:
            if bt.get("groupId") == target["id"]:
                rtids.add(bt["id"])
        rpids = set()
        for bt in all_bt:
            if bt["id"] in rtids and bt.get("projectId"):
                rpids.add(bt["projectId"])
        if target.get("projectId"):
            rpids.add(target["projectId"])

        tevents = [e for e in events if e.get("projectId") in rpids and ps <= e["date"] <= pe]
        if not tevents:
            continue

        is_grouped = any(bt.get("groupId") == target["id"] for bt in all_bt)
        grouped = {}
        for e in tevents:
            k = f"{e.get('managerId') or ''}|{e.get('projectId') or ''}|{e.get('taskId') or ''}|{e.get('note') or ''}"
            if k not in grouped:
                grouped[k] = {"managerId": e.get("managerId"), "projectId": e.get("projectId"),
                              "taskId": e.get("taskId"), "note": e.get("note"), "dates": [], "tp": 0}
            grouped[k]["dates"].append(e["date"])
            grouped[k]["tp"] += int(e.get("peopleCount") or 0)

        by_mgr = {}
        for g in grouped.values():
            mid = g["managerId"]
            if mid not in by_mgr:
                by_mgr[mid] = []
            by_mgr[mid].append(g)

        line_rows = []
        for mid, items in by_mgr.items():
            mn = managers_map.get(mid, {}).get("name") if mid else None
            if mn:
                line_rows.append({"type": "manager", "name": mn})
            for item in items:
                pn = projects_map.get(item["projectId"], {}).get("name") if is_grouped else None
                tn = tasks_map.get(item["taskId"], {}).get("name", "") if item["taskId"] else ""
                mo = (item["note"] or "").strip()
                parts = [x for x in [pn, tn, mo] if x]
                line_rows.append({
                    "type": "item", "label": "　".join(parts),
                    "dateStr": fmt_dates(item["dates"]),
                    "qty": item["tp"] if target.get("billingType") == "人工" else 1,
                    "unit": "人" if target.get("billingType") == "人工" else "式",
                    "unitPrice": target.get("unitPrice"),
                })

        is_sanei = target["name"] == SANEI
        is_combo = target.get("outputType") == "請求書兼明細書"
        tmpl = SANEI if is_sanei else ("請求書兼明細書" if is_combo else "請求書＋明細書")

        ws_src = wb_main[tmpl]
        ws_new = wb_main.copy_worksheet(ws_src)
        copy_sheet_styles(ws_src, ws_new)

        sname = target["name"][:31]
        base_name = sname
        n = 1
        existing = [s.title for s in wb_main.worksheets]
        while sname in existing:
            sname = f"{base_name[:28]}_{n}"
            n += 1
        ws_new.title = sname

        if not is_sanei:
            ws_new["A1"] = target["name"]
            if not is_combo:
                ws_new["L1"] = target["name"]
        ws_new["I4"] = f"{year}年"
        ws_new["J4"] = f"{period_month}月　{cd}日〆"
        if not is_combo:
            ws_new["T4"] = f"{year}年"
            ws_new["U4"] = f"{period_month}月　{cd}日〆"

        lc, dc, qc, uc, pc2 = ("A", "E", "F", "G", "H") if is_combo else ("L", "P", "Q", "R", "S")

        rn = 17
        for row in line_rows:
            if rn > 32:
                break
            if row["type"] == "manager" and row.get("name"):
                ws_new[f"{lc}{rn}"] = f"{row['name']}　様"
                rn += 1
            elif row["type"] == "item":
                ws_new[f"{lc}{rn}"] = row["label"]
                ws_new[f"{dc}{rn}"] = row["dateStr"]
                ws_new[f"{qc}{rn}"] = row["qty"]
                ws_new[f"{uc}{rn}"] = row["unit"]
                if row.get("unitPrice") is not None:
                    ws_new[f"{pc2}{rn}"] = row["unitPrice"]
                rn += 1

        if not is_combo:
            rn = 17
            for row in line_rows:
                if rn > 32:
                    break
                if row["type"] == "manager" and row.get("name"):
                    ws_new[f"A{rn}"] = f"{row['name']}　様"
                    rn += 1
                elif row["type"] == "item":
                    ws_new[f"A{rn}"] = row["label"]
                    ws_new[f"E{rn}"] = row["dateStr"]
                    ws_new[f"F{rn}"] = row["qty"]
                    ws_new[f"G{rn}"] = row["unit"]
                    if row.get("unitPrice") is not None:
                        ws_new[f"H{rn}"] = row["unitPrice"]
                    rn += 1

        added_sheets.append((sname, target["name"], target.get("outputType", ""), f"{cd}日〆"))

    for sname, tname, otype, cstr in added_sheets:
        cell = ws_idx.cell(row=index_row, column=1, value=tname)
        cell.hyperlink = f"#{sname}!A1"
        cell.font = Font(color="0000FF", underline="single")
        ws_idx.cell(row=index_row, column=2, value=otype)
        ws_idx.cell(row=index_row, column=3, value=cstr)
        index_row += 1

    added_names = {s[0] for s in added_sheets}
    for tmpl_name in ["請求書＋明細書", "請求書兼明細書", SANEI]:
        if tmpl_name in wb_main.sheetnames and tmpl_name not in added_names:
            wb_main[tmpl_name].sheet_state = "hidden"

    wb_main.move_sheet("INDEX", offset=-len(wb_main.sheetnames) + 1)

    out = BytesIO()
    wb_main.save(out)
    out.seek(0)
    return base64.b64encode(out.read()).decode()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(length))
        try:
            result = generate_invoice(body)
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"file": result}).encode())
        except Exception as e:
            import traceback
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e), "trace": traceback.format_exc()}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
